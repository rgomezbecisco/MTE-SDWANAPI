"""Excel report generation helpers for SD-WAN data."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from mte_sdwanapi.api import parse_feature_templates_for_device_template, tprint


def _get_default_feature_template_ids(feature_template_definitions):
    """Return a set of template IDs that are factory defaults."""
    default_ids = set()
    for feature_template in feature_template_definitions:
        if feature_template.get("factoryDefault"):
            template_id = feature_template.get("templateId")
            if template_id:
                default_ids.add(template_id)
    return default_ids


def _build_default_to_device_templates_map(device_template_definitions, default_feature_template_ids):
    """Return a mapping of default feature template IDs to device templates using them."""
    default_to_device_map = {default_id: [] for default_id in default_feature_template_ids}
    
    for device_template in device_template_definitions:
        device_template_id = device_template.get("templateId")
        device_template_name = device_template.get("templateName", "Unknown")
        
        # Extract all feature template IDs referenced by this device template
        feature_templates = parse_feature_templates_for_device_template(device_template)
        for feature_template in feature_templates:
            ft_id = feature_template.get("templateId")
            if ft_id in default_feature_template_ids:
                default_to_device_map[ft_id].append({
                    "device_template_id": device_template_id,
                    "device_template_name": device_template_name,
                })
    
    return default_to_device_map


def save_report_to_excel(
    destination_file,
    reachable_devices,
    device_template_definitions,
    feature_template_definitions,
    policy_definitions,
    attached_device_values,
):
    """Build and save the Excel report workbook."""

    if not reachable_devices:
        tprint("No reachable devices found. Creating report with empty device section.")

    workbook = Workbook()
    reachable_devices_worksheet = workbook.active
    reachable_devices_worksheet.title = "reachable devices"

    headers = list(reachable_devices[0].keys()) if reachable_devices else ["status", "details"]

    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = reachable_devices_worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    if reachable_devices:
        for row_idx, device in enumerate(reachable_devices, start=2):
            for col_idx, header in enumerate(headers, start=1):
                reachable_devices_worksheet.cell(
                    row=row_idx, column=col_idx, value=device.get(header)
                )
    else:
        reachable_devices_worksheet.cell(row=2, column=1, value="NO_REACHABLE_DEVICES")
        reachable_devices_worksheet.cell(
            row=2,
            column=2,
            value="No devices were reachable at report generation time.",
        )

    for col_idx, header in enumerate(headers, start=1):
        if reachable_devices:
            max_data_length = max(
                len(str(device.get(header, ""))) for device in reachable_devices
            )
        elif header == "status":
            max_data_length = len("NO_REACHABLE_DEVICES")
        else:
            max_data_length = len("No devices were reachable at report generation time.")
        reachable_devices_worksheet.column_dimensions[
            reachable_devices_worksheet.cell(row=1, column=col_idx).column_letter
        ].width = max(len(header), max_data_length) + 2

    reachable_devices_worksheet.freeze_panes = "A2"

    feature_template_lookup = {}
    for feature_template_definition in feature_template_definitions:
        template_id = feature_template_definition.get("templateId")
        if template_id:
            feature_template_lookup[template_id] = feature_template_definition

    default_feature_template_ids = _get_default_feature_template_ids(feature_template_definitions)
    
    # Create summary sheet for default feature templates
    if default_feature_template_ids:
        summary_worksheet = workbook.create_sheet(title="default templates summary", index=1)
        
        summary_headers = ["Default Feature Template", "Template ID", "Device Templates Using It"]
        for col_idx, header in enumerate(summary_headers, start=1):
            cell = summary_worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        default_to_device_map = _build_default_to_device_templates_map(
            device_template_definitions, default_feature_template_ids
        )
        
        row_idx = 2
        for default_ft_id in sorted(default_feature_template_ids):
            default_ft_def = feature_template_lookup.get(default_ft_id, {})
            template_name = default_ft_def.get("templateName", "Unknown")
            device_templates_using = default_to_device_map.get(default_ft_id, [])
            
            device_template_names = ", ".join(
                [dt.get("device_template_name", "Unknown") for dt in device_templates_using]
            )
            
            summary_worksheet.cell(row=row_idx, column=1, value=template_name)
            summary_worksheet.cell(row=row_idx, column=2, value=default_ft_id)
            summary_worksheet.cell(row=row_idx, column=3, value=device_template_names)
            
            row_idx += 1
        
        summary_worksheet.column_dimensions["A"].width = 35
        summary_worksheet.column_dimensions["B"].width = 40
        summary_worksheet.column_dimensions["C"].width = 60

    policy_lookup = {}
    for policy_definition in policy_definitions:
        policy_id = policy_definition.get("policyId")
        if policy_id:
            policy_lookup[policy_id] = policy_definition

    attached_values_lookup = {}
    for template_data in attached_device_values:
        template_id = template_data.get("templateId")
        if template_id:
            attached_values_lookup[template_id] = template_data.get(
                "attachedValues", []
            )

    used_sheet_names = {reachable_devices_worksheet.title}
    for idx, device_template in enumerate(device_template_definitions, start=1):
        template_name = device_template.get("templateName") or f"device_template_{idx}"
        sheet_name = template_name[:31].replace("*","")
        if sheet_name in used_sheet_names:
            sheet_name = f"Template {idx}"[:31]
        used_sheet_names.add(sheet_name)

        worksheet = workbook.create_sheet(title=sheet_name)

        summary_headers = ["Field", "Value"]
        for col_idx, header in enumerate(summary_headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        summary_rows = [
            ("Template Name", device_template.get("templateName", "")),
            ("Template ID", device_template.get("templateId", "")),
            (
                "Template Description",
                device_template.get("templateDescription", ""),
            ),
            ("Policy ID", device_template.get("policyId", "")),
            (
                "Policy Name",
                policy_lookup.get(device_template.get("policyId"), {}).get(
                    "policyName", "NOT_FOUND"
                ),
            ),
        ]

        for row_idx, row_data in enumerate(summary_rows, start=2):
            worksheet.cell(row=row_idx, column=1, value=row_data[0])
            worksheet.cell(row=row_idx, column=2, value=row_data[1])

        # Keep tables aligned by using a fixed starting row on each sheet.
        feature_table_header_row = 8
        feature_headers = [
            "templateType",
            "templateName",
            "templateDescription",
            "templateId",
        ]
        for col_idx, header in enumerate(feature_headers, start=1):
            cell = worksheet.cell(row=feature_table_header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        template_feature_templates = parse_feature_templates_for_device_template(
            device_template
        )
        red_font = Font(color="FF0000", bold=True)
        for row_idx, template_feature in enumerate(
            template_feature_templates, start=feature_table_header_row + 1
        ):
            template_id = template_feature.get("templateId")
            feature_definition = feature_template_lookup.get(template_id, {})
            is_default = template_id in default_feature_template_ids
            font_style = red_font if is_default else Font()
            
            worksheet.cell(
                row=row_idx,
                column=1,
                value=template_feature.get("templateType", ""),
            )
            worksheet.cell(row=row_idx, column=1).font = font_style
            
            worksheet.cell(
                row=row_idx,
                column=2,
                value=feature_definition.get("templateName", "NOT_FOUND"),
            )
            worksheet.cell(row=row_idx, column=2).font = font_style
            
            worksheet.cell(
                row=row_idx,
                column=3,
                value=feature_definition.get("templateDescription", "NOT_FOUND"),
            )
            worksheet.cell(row=row_idx, column=3).font = font_style
            
            worksheet.cell(row=row_idx, column=4, value=template_id)
            worksheet.cell(row=row_idx, column=4).font = font_style

        # Start attached-variable columns to the right of the feature table.
        attached_table_start_col = 9
        attached_table_header_row = 8
        attached_rows = attached_values_lookup.get(device_template.get("templateId"), [])

        preferred_attached_headers = [
            "csv-deviceIP",
            "csv-deviceId",
            "csv-host-name",
        ]

        discovered_headers = []
        for row_data in attached_rows:
            for key in row_data.keys():
                if key not in discovered_headers:
                    discovered_headers.append(key)

        attached_headers = [
            header for header in preferred_attached_headers if header in discovered_headers
        ]
        attached_headers.extend(
            sorted([header for header in discovered_headers if header not in attached_headers])
        )

        if not attached_headers:
            attached_headers = ["attached-values"]

        for col_offset, header in enumerate(attached_headers):
            cell = worksheet.cell(
                row=attached_table_header_row,
                column=attached_table_start_col + col_offset,
                value=header,
            )
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        if attached_rows:
            for row_idx, row_data in enumerate(attached_rows, start=attached_table_header_row + 1):
                for col_offset, header in enumerate(attached_headers):
                    value = row_data.get(header)
                    if isinstance(value, list):
                        value = ", ".join(str(item) for item in value)
                    worksheet.cell(
                        row=row_idx,
                        column=attached_table_start_col + col_offset,
                        value=value,
                    )
        else:
            worksheet.cell(
                row=attached_table_header_row + 1,
                column=attached_table_start_col,
                value="No attached devices/variables",
            )

        worksheet.column_dimensions["A"].width = 24
        worksheet.column_dimensions["B"].width = 50
        worksheet.column_dimensions["C"].width = 60
        worksheet.column_dimensions["D"].width = 40
        for col_offset, header in enumerate(attached_headers):
            col_letter = worksheet.cell(
                row=attached_table_header_row,
                column=attached_table_start_col + col_offset,
            ).column_letter
            if attached_rows:
                max_data_len = max(
                    len(
                        str(
                            (
                                row_data.get(header)
                                if not isinstance(row_data.get(header), list)
                                else ", ".join(
                                    str(item) for item in row_data.get(header, [])
                                )
                            )
                            or ""
                        )
                    )
                    for row_data in attached_rows
                )
            else:
                max_data_len = len("No attached devices/variables") if col_offset == 0 else 0
            worksheet.column_dimensions[col_letter].width = max(len(header), max_data_len) + 2

    workbook.save(destination_file)
    tprint(f"{destination_file} saved!")
